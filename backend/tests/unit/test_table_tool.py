"""Tests for the SQL query tool over tabular documents."""

from pathlib import Path

import openpyxl
import pytest

from hivegent.tools.base import ToolRetry
from hivegent.tools.table import QueryTableTool
from tests.helpers import returned


def _sales_dir(tmp_path: Path) -> Path:
    """Write a 50-row table whose EU rows sum to 625 and US rows to 600."""
    rows = "\n".join(f"{'EU' if i % 2 else 'US'},{i}" for i in range(50))
    (tmp_path / "sales.csv").write_text(f"region,amount\n{rows}")

    return tmp_path


def _workbook_dir(tmp_path: Path) -> Path:
    book = openpyxl.Workbook()
    first = book.active
    assert first is not None
    first.title = "Q1"
    first.append(["item", "qty"])
    first.append(["bolt", 7])
    second = book.create_sheet("Q2")
    second.append(["item", "qty"])
    second.append(["nut", 3])
    book.save(tmp_path / "book.xlsx")

    return tmp_path


class TestQueryTableTool:
    """SQL over the original file, so a large table costs a small result."""

    async def test_schema_mode_reports_columns_and_row_count(
        self, tmp_path: Path
    ) -> None:
        tool = QueryTableTool(paths=_sales_dir(tmp_path))
        out = await returned(tool("sales.csv"))

        assert out.data.columns == ("region", "amount")
        assert out.data.total_rows == 50
        assert len(out.data.rows) == tool.preview_rows
        assert out.formatted is not None
        assert "amount: Int64" in out.formatted

    async def test_query_aggregates_instead_of_returning_rows(
        self, tmp_path: Path
    ) -> None:
        tool = QueryTableTool(paths=_sales_dir(tmp_path))
        out = await returned(
            tool(
                "sales.csv",
                "SELECT region, SUM(amount) AS total FROM t GROUP BY region "
                "ORDER BY region",
            )
        )

        assert out.data.rows == (("EU", "625"), ("US", "600"))

    async def test_row_cap_is_named_in_the_output(self, tmp_path: Path) -> None:
        tool = QueryTableTool(paths=_sales_dir(tmp_path), max_rows=10)
        out = await returned(tool("sales.csv", "SELECT * FROM t"))

        assert out.data.truncated
        assert len(out.data.rows) == 10
        assert out.formatted is not None
        assert "10 rows returned" in out.formatted

    async def test_model_can_request_more_than_default_row_limit(
        self, tmp_path: Path
    ) -> None:
        rows = "\n".join(str(index) for index in range(150))
        (tmp_path / "rows.csv").write_text(f"value\n{rows}")
        tool = QueryTableTool(paths=tmp_path)

        out = await returned(tool("rows.csv", "SELECT * FROM t", row_limit=150))

        assert len(out.data.rows) == 150
        assert not out.data.truncated

    async def test_display_budget_binds_the_text_and_not_the_rows(
        self, tmp_path: Path
    ) -> None:
        # The budget once ended the row loop, so rows past it never reached
        # `rows` either — and a redirect then wrote a partial table under a
        # receipt that read like a whole one.
        tool = QueryTableTool(paths=_sales_dir(tmp_path), max_formatted_chars=100)

        out = await returned(tool("sales.csv", "SELECT * FROM t", row_limit=50))

        assert len(out.data.rows) == 50
        assert not out.data.truncated
        assert out.formatted is not None
        assert out.formatted.count("\n|") < 50
        assert "all 50 rows are in the result" in out.formatted

    async def test_legacy_encoding_is_decoded_and_reported(
        self, tmp_path: Path
    ) -> None:
        # A lazy scan only fails once it reaches the offending bytes, so the
        # retry has to survive them sitting anywhere in the file, not just in
        # the header a cheap probe would have seen.
        filler = "\n".join(f"n{i},c{i}" for i in range(200))
        (tmp_path / "legacy.csv").write_bytes(
            f"name,city\n{filler}\nGrüße,Köln\n".encode("cp1252")
        )
        tool = QueryTableTool(paths=tmp_path)
        out = await returned(tool("legacy.csv", "SELECT * FROM t WHERE city = 'Köln'"))

        assert out.data.tables[0].source_encoding == "cp1252"
        assert out.data.rows == (("Grüße", "Köln"),)

    async def test_worksheet_is_selectable_and_the_others_are_named(
        self, tmp_path: Path
    ) -> None:
        tool = QueryTableTool(paths=_workbook_dir(tmp_path))
        out = await returned(tool("book.xlsx", sheet="Q2"))

        assert out.data.tables[0].sheet == "Q2"
        assert out.data.tables[0].sheets == ("Q1", "Q2")
        assert out.data.rows == (("nut", "3"),)

        with pytest.raises(ToolRetry, match="Q1, Q2"):
            await tool("book.xlsx", sheet="Q9")

    async def test_column_cut_is_named_and_leaves_the_data_channel_whole(
        self, tmp_path: Path
    ) -> None:
        # The cut binds the rendering only: the frontend still gets every
        # column, the way a read keeps its lines untruncated in `content`.
        header = ",".join(f"c{i}" for i in range(12))
        values = ",".join(str(i) for i in range(12))
        (tmp_path / "wide.csv").write_text(f"{header}\n{values}")
        tool = QueryTableTool(paths=tmp_path, max_columns=4)
        out = await returned(tool("wide.csv", "SELECT * FROM t"))

        assert len(out.data.columns) == 12
        assert out.formatted is not None
        assert "4 of 12 columns shown" in out.formatted
        assert "c11" not in out.formatted

    async def test_text_columns_are_retyped_when_every_value_parses(
        self, tmp_path: Path
    ) -> None:
        # A spreadsheet hands out numbers and dates as text, which is what
        # otherwise turns a plain SUM or a date comparison into a dtype error.
        book = openpyxl.Workbook()
        sheet = book.active
        assert sheet is not None
        sheet.append(["zip", "amount", "day"])
        sheet.append(["01067", "10.5", "2024-01-02"])
        sheet.append(["10115", "20", "2024-02-03"])
        book.save(tmp_path / "typed.xlsx")
        tool = QueryTableTool(paths=tmp_path)

        out = await returned(
            tool(
                "typed.xlsx", "SELECT SUM(amount) AS s FROM t WHERE day > '2024-01-15'"
            )
        )
        assert out.data.rows == (("20.0",),)

        # A zero-padded value is an identifier, so the column stays text.
        schema = await returned(tool("typed.xlsx"))
        assert schema.data.dtypes == ("String", "Float64", "Date")
        assert [column.name for column in schema.data.text_columns] == [
            "amount",
            "day",
        ]

    async def test_mixed_column_is_reported_before_a_query_fails_over_it(
        self, tmp_path: Path
    ) -> None:
        # One "N/A" keeps the column text; naming the share that parses is
        # what lets the first query wrap it in TRY_CAST rather than the second.
        rows = "\n".join(f"r{i},{i if i != 2 else 'N/A'}" for i in range(4))
        (tmp_path / "mixed.csv").write_text(f"name,val\n{rows}")
        tool = QueryTableTool(paths=tmp_path)
        out = await returned(tool("mixed.csv"))

        assert out.data.text_columns[0].name == "val"
        assert (out.data.text_columns[0].parsed, out.data.text_columns[0].total) == (
            3,
            4,
        )
        assert out.formatted is not None
        # The count says the column is mixed; the value says what with, which
        # is the difference between a label row, a censored reading, and a typo.
        assert 'val (3 of 4 parse as Int64; unparsed: "N/A")' in out.formatted

        with pytest.raises(ToolRetry, match="TRY_CAST"):
            await tool("mixed.csv", "SELECT SUM(val) FROM t")

    async def test_mixed_column_is_reported_on_a_query_that_names_it(
        self, tmp_path: Path
    ) -> None:
        # The call that most needs to hear it is the one that named the column:
        # a run that opens with a SELECT never asks for the schema afterwards.
        rows = "\n".join(f"r{i},{i if i != 2 else 'N/A'}" for i in range(4))
        (tmp_path / "mixed.csv").write_text(f"name,val\n{rows}")
        tool = QueryTableTool(paths=tmp_path)

        out = await returned(tool("mixed.csv", "SELECT name, val FROM t"))

        assert out.formatted is not None
        assert 'val (3 of 4 parse as Int64; unparsed: "N/A")' in out.formatted

    async def test_select_star_hears_about_every_column(self, tmp_path: Path) -> None:
        # `SELECT *` spells no column name, so scoping on the query text alone
        # dropped every warning from the one query that asks for everything.
        rows = "\n".join(f"r{i},{i if i != 2 else 'N/A'}" for i in range(4))
        (tmp_path / "mixed.csv").write_text(f"name,val\n{rows}")
        tool = QueryTableTool(paths=tmp_path)

        out = await returned(tool("mixed.csv", "SELECT * FROM t"))

        assert out.formatted is not None
        assert 'val (3 of 4 parse as Int64; unparsed: "N/A")' in out.formatted

    async def test_a_query_hears_nothing_about_columns_it_did_not_name(
        self, tmp_path: Path
    ) -> None:
        rows = "\n".join(f"r{i},{i if i != 2 else 'N/A'}" for i in range(4))
        (tmp_path / "mixed.csv").write_text(f"name,val\n{rows}")
        tool = QueryTableTool(paths=tmp_path)

        out = await returned(tool("mixed.csv", "SELECT name FROM t"))

        assert out.formatted is not None
        assert "TRY_CAST" not in out.formatted

    async def test_the_worst_shortfall_leads_and_the_cap_is_named(
        self, tmp_path: Path
    ) -> None:
        # Cutting by position once hid the only column a query was about
        # behind nine it was not, and said nothing about having cut anything.
        header = ",".join(f"c{i}" for i in range(4))
        rows = "\n".join(
            ",".join("x" if i <= column else str(i) for column in range(4))
            for i in range(4)
        )
        (tmp_path / "wide.csv").write_text(f"{header}\n{rows}")
        tool = QueryTableTool(paths=tmp_path, max_named_columns=2)

        formatted = (await returned(tool("wide.csv"))).formatted

        assert formatted is not None
        mixed = formatted.split("mixed text, wrap in TRY_CAST to compare: ")[1]
        assert mixed.startswith("c2 (1 of 4")
        assert "and 1 more column" in mixed

    async def test_an_unparsable_value_is_named_and_handed_back(
        self, tmp_path: Path
    ) -> None:
        # `<100` is a lab declining to put a number on a sample it measured,
        # and every substitution for one moves every total. Naming the value
        # is what lets that be asked about; no rule for `<` is needed to see it.
        (tmp_path / "lab.csv").write_text("date,csb\n2023-01-01,153\n2023-01-02,<100")
        tool = QueryTableTool(paths=tmp_path)

        out = await returned(tool("lab.csv", "SELECT date, csb FROM t"))

        assert out.data.text_columns[0].unparsed == ("<100",)
        assert out.formatted is not None
        assert 'csb (1 of 2 parse as Int64; unparsed: "<100")' in out.formatted
        assert "ask the user" in out.formatted

    async def test_a_clean_table_says_nothing_about_its_columns(
        self, tmp_path: Path
    ) -> None:
        tool = QueryTableTool(paths=_sales_dir(tmp_path))

        formatted = (await returned(tool("sales.csv"))).formatted

        assert formatted is not None
        assert "unparsed" not in formatted
        assert "ask the user" not in formatted


    async def test_non_tabular_file_is_refused(self, tmp_path: Path) -> None:
        (tmp_path / "notes.md").write_text("plain prose")
        tool = QueryTableTool(paths=tmp_path)

        with pytest.raises(ToolRetry, match="not a tabular file"):
            await tool("notes.md")

    async def test_bad_query_names_the_valid_columns(self, tmp_path: Path) -> None:
        # Polars' own message carries the retry signal; the tool only has to
        # let it through rather than flatten it into "query failed".
        tool = QueryTableTool(paths=_sales_dir(tmp_path))

        with pytest.raises(ToolRetry, match="region"):
            await tool("sales.csv", "SELECT nope FROM t")


def _quoted_dir(tmp_path: Path) -> Path:
    """A table whose first column name is not a bare SQL identifier."""
    (tmp_path / "t.csv").write_text("Zulaufmenge (D),amount\n1,2\n")

    return tmp_path


def _exported_dir(tmp_path: Path) -> Path:
    """A sheet whose header spans two rows, as an export routinely does."""
    book = openpyxl.Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.append(["Datum", None, "Zulaufmenge (D)"])
    sheet.append(["Einstellung", "Meiches", "Zaehlwert"])
    sheet.append(["2023-01-01", "0.4", "9362"])
    book.save(tmp_path / "export.xlsx")

    return tmp_path


class TestMultipleTables:
    """A query spans every file it was given, so a join needs only the SQL."""

    def _both(self, tmp_path: Path) -> Path:
        (tmp_path / "sales.csv").write_text("id,amount\n1,10\n2,20\n")
        (tmp_path / "regions.csv").write_text("id,region\n1,EU\n2,US\n")

        return tmp_path

    async def test_two_tables_join_under_positional_names(
        self, tmp_path: Path
    ) -> None:
        tool = QueryTableTool(paths=self._both(tmp_path))

        out = await returned(
            tool(
                ["sales.csv", "regions.csv"],
                "SELECT t2.region, SUM(t.amount) AS total FROM t "
                "JOIN t2 ON t.id = t2.id GROUP BY t2.region ORDER BY t2.region",
            )
        )

        assert out.data.rows == (("EU", "10"), ("US", "20"))
        assert [table.name for table in out.data.tables] == ["t", "t2"]

    async def test_each_table_is_named_with_what_it_is_called(
        self, tmp_path: Path
    ) -> None:
        # A list of paths says nothing about which took which name, and a join
        # cannot be written without knowing.
        tool = QueryTableTool(paths=self._both(tmp_path))

        out = await returned(tool(["sales.csv", "regions.csv"], "SELECT * FROM t2"))

        assert out.formatted is not None
        assert "t: sales.csv" in out.formatted
        assert "t2: regions.csv" in out.formatted

    async def test_show_tables_answers_from_the_query_context(
        self, tmp_path: Path
    ) -> None:
        tool = QueryTableTool(paths=self._both(tmp_path))

        out = await returned(tool(["sales.csv", "regions.csv"], "SHOW TABLES"))

        assert out.data.rows == (("t",), ("t2",))

    async def test_a_mixed_column_is_qualified_by_its_table(
        self, tmp_path: Path
    ) -> None:
        # A bare name is what a single-table query types; `t2.val` is what a
        # join has to, so the spelling follows the shape of the call.
        (tmp_path / "a.csv").write_text("val\n1\n")
        (tmp_path / "b.csv").write_text("val\n1\nN/A\n")
        tool = QueryTableTool(paths=tmp_path)

        out = await returned(tool(["a.csv", "b.csv"], "SELECT t2.val FROM t2"))

        assert out.formatted is not None
        assert 't2.val (1 of 2 parse as Int64; unparsed: "N/A")' in out.formatted

    async def test_a_single_path_still_takes_a_bare_string(
        self, tmp_path: Path
    ) -> None:
        tool = QueryTableTool(paths=_sales_dir(tmp_path))

        out = await returned(tool("sales.csv", "SELECT COUNT(*) AS n FROM t"))

        assert out.data.rows == (("50",),)


class TestColumnSpelling:
    """The schema listing is the only place a model sees a column name."""

    async def test_a_name_that_is_not_an_identifier_is_listed_quoted(
        self, tmp_path: Path
    ) -> None:
        tool = QueryTableTool(paths=_quoted_dir(tmp_path))

        formatted = (await returned(tool("t.csv"))).formatted
        assert formatted is not None
        assert '"Zulaufmenge (D)": Int64' in formatted
        assert "\namount: Int64" in formatted

    async def test_a_parse_error_names_the_quoting_rule(self, tmp_path: Path) -> None:
        tool = QueryTableTool(paths=_quoted_dir(tmp_path))

        with pytest.raises(ToolRetry, match="double-quoted") as exc:
            await tool("t.csv", "SELECT Zulaufmenge (D) FROM t")

        assert '"Zulaufmenge (D)"' in str(exc.value)

    async def test_the_quoted_spelling_is_the_one_that_runs(
        self, tmp_path: Path
    ) -> None:
        tool = QueryTableTool(paths=_quoted_dir(tmp_path))

        out = await returned(tool("t.csv", 'SELECT "Zulaufmenge (D)" FROM t'))
        assert out.data.rows == (("1",),)

    async def test_a_spilled_header_row_shows_up_as_its_own_labels(
        self, tmp_path: Path
    ) -> None:
        # One label row is enough to hold every numeric column in a sheet as
        # text.  Nothing here rules on whether row 1 is labels or data — only
        # the sheet's author knows — but naming the value in each column is
        # what makes it visible, and the same label repeated across columns
        # is what says the header spilled a row rather than one cell being odd.
        tool = QueryTableTool(paths=_exported_dir(tmp_path))

        out = await returned(tool("export.xlsx"))

        assert out.formatted is not None
        assert 'unparsed: "Einstellung"' in out.formatted
        assert 'unparsed: "Zaehlwert"' in out.formatted

    async def test_a_header_spanning_three_rows_reads_the_same_way(
        self, tmp_path: Path
    ) -> None:
        # The reason this is one mechanism and not a rule per shape: a rule
        # keyed on "row 1 holds the column's only bad value" sees nothing here.
        book = openpyxl.Workbook()
        sheet = book.active
        assert sheet is not None
        sheet.append(["Datum", "Zulauf"])
        sheet.append(["Einstellung", "Zaehlwert"])
        sheet.append(["", "m3/d"])
        sheet.append(["2023-01-01", "9362"])
        book.save(tmp_path / "deep.xlsx")
        tool = QueryTableTool(paths=tmp_path)

        out = await returned(tool("deep.xlsx"))

        assert out.formatted is not None
        assert 'unparsed: "Zaehlwert", "m3/d"' in out.formatted

    async def test_a_clean_header_says_nothing_about_one(self, tmp_path: Path) -> None:
        tool = QueryTableTool(paths=_sales_dir(tmp_path))

        formatted = (await returned(tool("sales.csv"))).formatted
        assert formatted is not None
        assert "no header name" not in formatted
